import openpyxl

fpath = r'C:\Users\yongug\STARTCODING\STARTCODING_CRAWLING\02_파이썬엑셀다루기\참가자_data.xlsx'

#1) 액셀 불러오기


wb = openpyxl.load_workbook(fpath)

#2) 시트 불러오기
ws = wb['오징어게임']


#3) 데이터 수정하기
ws['A3'] = '456'
ws['B3'] = '성기훈'

wb.save(fpath)